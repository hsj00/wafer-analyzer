# modules/report.py
# Excel 보고서 자동 생성 모듈
# 통계표 + 웨이퍼 맵 이미지 + 원시 데이터를 xlsx로 다운로드
#
# pip install openpyxl kaleido
#
# =============================================================================
# [설계 결정 근거]
# =============================================================================
#
# ① @st.cache_data 적용 불가 함수들의 실행 시점 제어
#    go.Figure는 mutable → hash() 불가 → @st.cache_data에 Figure를 인자로 전달하면
#    TypeError 발생. generate_excel_report, safe_fig_to_png 모두 Figure를 받으므로
#    캐시 적용 불가.
#    → 두 함수 모두 일반 함수로 구현.
#    → render_report_tab에서 st.button 조건부 실행으로 클릭 시에만 호출.
#    → 결과(xl_bytes)를 session_state["rep_bytes"]에 저장해 다음 rerun에도 유지.
#
# ② kaleido 탐지: 모듈 로딩 시 1회 probe
#    단순 import kaleido 체크보다 실제 변환 가능 여부를 검사하는 것이 더 정확.
#    매우 작은 더미 Figure로 fig.to_image()를 시도해 _KALEIDO_OK 플래그 설정.
#    → safe_fig_to_png 내부에서 매번 try/except 대신 플래그로 조기 반환.
#    → 모듈 로딩 시 probe 실패는 조용히 흡수 (앱 시작 방해 없음).
#
# ③ openpyxl BytesIO 이미지 삽입 — 수명 관리
#    XLImage(BytesIO) 생성 후 wb.save() 시점에 BytesIO 내용을 실제로 읽음.
#    → BytesIO 객체가 GC되면 XLImage 내부 참조 끊김 → 저장 실패 가능성.
#    → _img_refs 리스트로 generate_excel_report 스코프 동안 강제 유지.
#    → 리스트가 함수 반환 후 GC → 안전하게 해제.
#
# ④ 이미지 배치: 2열 × 2행 그리드
#    A1  → Heatmap    / I1  → Contour
#    A31 → Line Scan  / I31 → 3D Surface
#    행 오프셋 30 = 이미지 높이 350px ÷ (엑셀 행 높이 약 15px) + 여유 1행
#    열 오프셋 8 = 이미지 폭 400px ÷ (엑셀 열 너비 약 8px 기준)
#
# ⑤ 원시 데이터 5000행 제한
#    대용량 파일(수만 행)의 전체 포함 시 xlsx 파일이 수십 MB로 팽창.
#    → head(5000)으로 제한 + 초과 시 경고 텍스트 삽입.
#    → 사용자가 render_report_tab에서 nrows 슬라이더로 조정 가능.
# =============================================================================

# ── 표준 라이브러리 ─────────────────────────────────────────────────────────
import io
import logging
import os
from datetime import datetime

# ── 외부 라이브러리 ─────────────────────────────────────────────────────────
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
# openpyxl: Excel 파일 생성 (pip install openpyxl)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 로거 설정 ────────────────────────────────────────────────────────────────
_logger = logging.getLogger(__name__)


# =============================================================================
# session_state 키 상수 (prefix: "rep_")
# =============================================================================
# 기존 키: data_folder, datasets, _s_display 등
# 다른 모듈 키: mp_*, def_*, gpc_* (충돌 없음)
_SS_BYTES       = "rep_bytes"        # 생성된 xlsx bytes (다운로드용)
_SS_GENERATING  = "rep_generating"   # 중복 클릭 방지 플래그 (bool)
_SS_INC_MAPS    = "rep_inc_maps"     # 웨이퍼 맵 이미지 포함 여부 (bool)
_SS_INC_RAW     = "rep_inc_raw"      # 원시 데이터 포함 여부 (bool)
_SS_INC_GPC     = "rep_inc_gpc"      # GPC 분석 포함 여부 (bool)
_SS_NROWS       = "rep_nrows"        # 원시 데이터 최대 행 수 (int)


# =============================================================================
# openpyxl 스타일 상수
# =============================================================================
_HEADER_FILL = PatternFill("solid", start_color="1A6BBF")   # 파란 헤더 배경
_ALT_FILL    = PatternFill("solid", start_color="E8F0FE")   # 교번 행 배경 (연파랑)
_WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")   # 흰색 행 배경

_BORDER_SIDE = Side(style="thin", color="C0C0C0")
_THIN_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE,  bottom=_BORDER_SIDE,
)

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_BODY_FONT   = Font(name="Arial", size=10)
_BOLD_FONT   = Font(name="Arial", bold=True, size=10)
_TITLE_FONT  = Font(name="Arial", bold=True, size=13, color="1A6BBF")
_META_FONT   = Font(name="Arial", size=9, italic=True, color="888888")
_WARN_FONT   = Font(name="Arial", size=9, italic=True, color="CC0000")

_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
_NUM_FORMAT   = "#,##0.0000"   # 숫자 소수점 4자리 형식


# =============================================================================
# kaleido 가용성 탐지 (모듈 로딩 시 1회 실행)
# =============================================================================

def _probe_kaleido() -> bool:
    """
    실제 변환 시도로 kaleido 가용성 확인.

    [import kaleido 체크보다 정확한 이유]
    kaleido가 설치됐더라도 실행 파일(kaleido binary)이 없거나
    버전 불일치이면 fig.to_image()가 RuntimeError를 발생시킴.
    → 실제 변환을 1회 시도해 _KALEIDO_OK 플래그 설정.

    [더미 Figure 크기]
    width=50, height=50: 최소 크기로 탐지 시간 최소화 (<0.5초 기대).
    모듈 로딩 시 1회만 실행되므로 사용자 체감 지연 없음.
    """
    try:
        dummy = go.Figure(go.Scatter(x=[1], y=[1]))
        dummy.to_image(format="png", width=50, height=50)
        return True
    except Exception:
        # kaleido 미설치, 바이너리 없음, 버전 불일치 등 모든 예외
        return False


# 모듈 로딩 시 1회 탐지 → 전역 플래그로 캐싱
# 이 플래그로 safe_fig_to_png 내부 불필요한 변환 시도 방지
try:
    _KALEIDO_OK: bool = _probe_kaleido()
except Exception:
    _KALEIDO_OK = False


# =============================================================================
# [함수 1] safe_fig_to_png
# =============================================================================

def safe_fig_to_png(
    fig: go.Figure,
    width: int = 600,
    height: int = 500,
) -> bytes | None:
    """
    go.Figure → PNG bytes 변환 (kaleido 없으면 None 반환).

    [@st.cache_data 미적용 이유]
    go.Figure는 mutable → hash() 불가 → @st.cache_data에 전달하면 TypeError.
    → 일반 함수로 구현. 버튼 클릭 시에만 호출되므로 성능 영향 없음.

    [kaleido 처리 전략]
    모듈 로딩 시 _probe_kaleido()로 _KALEIDO_OK 플래그 설정.
    → _KALEIDO_OK=False이면 변환 시도 없이 즉시 None 반환 (빠름).
    → _KALEIDO_OK=True이더라도 개별 변환 실패(메모리, 타임아웃 등)는
      try/except로 None 반환.

    인자:
        fig   : 변환할 Plotly Figure
        width : 출력 PNG 가로 픽셀 (기본 600)
        height: 출력 PNG 세로 픽셀 (기본 500)

    반환:
        bytes : PNG 이미지 데이터
        None  : kaleido 없음 또는 변환 실패
    """
    # kaleido 미설치 시 빠른 경로 반환 (변환 시도 자체를 건너뜀)
    if not _KALEIDO_OK:
        return None

    try:
        return fig.to_image(format="png", width=width, height=height)

    except ImportError:
        # kaleido가 탐지됐지만 런타임에 import 실패하는 엣지 케이스
        _logger.warning("kaleido ImportError: PNG 변환 실패")
        return None

    except Exception as e:
        # 메모리 부족, 타임아웃, Figure가 너무 복잡한 경우 등
        _logger.warning(f"PNG 변환 실패: {type(e).__name__}: {e}")
        return None


# =============================================================================
# openpyxl 내부 헬퍼 함수들
# =============================================================================

def _style_header_row(ws, row: int, n_cols: int, start_col: int = 1) -> None:
    """지정 행을 헤더 스타일(파란 배경, 흰 볼드 폰트, 가운데 정렬)로 설정."""
    for col in range(start_col, start_col + n_cols):
        cell            = ws.cell(row=row, column=col)
        cell.fill       = _HEADER_FILL
        cell.font       = _HEADER_FONT
        cell.border     = _THIN_BORDER
        cell.alignment  = _CENTER_ALIGN


def _style_data_rows(
    ws,
    start_row: int,
    end_row: int,
    n_cols: int,
    start_col: int = 1,
    first_col_left: bool = True,   # 첫 번째 컬럼 왼쪽 정렬 여부
) -> None:
    """
    데이터 행에 교번 배경색(짝수행=연파랑, 홀수행=흰색) + 테두리 적용.

    first_col_left=True: 첫 번째 컬럼은 왼쪽 정렬 (항목명/레이블 컬럼)
    나머지 컬럼: 가운데 정렬 (숫자 값 컬럼)
    """
    for row in range(start_row, end_row + 1):
        use_alt = (row % 2 == 0)   # 짝수 행에 교번 배경색 적용
        for col_offset in range(n_cols):
            col  = start_col + col_offset
            cell = ws.cell(row=row, column=col)
            cell.fill      = _ALT_FILL if use_alt else _WHITE_FILL
            cell.font      = _BODY_FONT
            cell.border    = _THIN_BORDER
            cell.alignment = (
                _LEFT_ALIGN if (first_col_left and col_offset == 0)
                else _CENTER_ALIGN
            )


def _set_number_format(ws, start_row: int, end_row: int,
                        col: int) -> None:
    """지정 컬럼의 숫자 셀에 소수점 4자리 형식 적용."""
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = _NUM_FORMAT


def _auto_col_width(ws, padding: int = 3, max_width: int = 50) -> None:
    """모든 컬럼 너비를 내용 최대 길이 + 여백으로 자동 조정."""
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=8,
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + padding, max_width)


# =============================================================================
# [함수 2] write_stats_sheet
# =============================================================================

def write_stats_sheet(
    ws,
    stats: dict,
    filename: str,
    measured_at: str,
) -> None:
    """
    통계 테이블을 openpyxl 워크시트에 작성.

    [시트 레이아웃]
    Row 1: 보고서 제목 (A1:B1 병합, 파란 폰트)
    Row 2: 파일명 + 생성일시 (메타 정보)
    Row 3: (빈 행)
    Row 4: 헤더 ("항목", "값") — 파란 배경 + 흰 볼드
    Row 5~: 통계 항목별 데이터 — 교번 배경 + 테두리
             숫자 컬럼: 소수점 4자리 형식

    인자:
        ws          : openpyxl 워크시트 객체
        stats       : calculate_stats() 반환 딕셔너리
        filename    : 원본 파일명 (헤더 메타 정보)
        measured_at : 분석 날짜/시간 문자열 (예: "2024-01-15 14:30")
    """
    # ── 제목 및 메타 정보 ─────────────────────────────────────────────────────
    ws["A1"] = "웨이퍼 맵 분석 통계 보고서"
    ws["A1"].font      = _TITLE_FONT
    ws["A1"].alignment = _CENTER_ALIGN
    ws.merge_cells("A1:B1")

    ws["A2"] = f"파일: {filename}    생성: {measured_at}"
    ws["A2"].font      = _META_FONT
    ws["A2"].alignment = _LEFT_ALIGN
    ws.merge_cells("A2:B2")

    # ── 헤더 행 ───────────────────────────────────────────────────────────────
    headers = ["항목", "값"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col_idx, value=h)
    _style_header_row(ws, row=4, n_cols=len(headers))

    # ── 통계 데이터 행 ────────────────────────────────────────────────────────
    for row_offset, (key, val) in enumerate(stats.items()):
        row = 5 + row_offset
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=val)

    end_row = 4 + len(stats)
    _style_data_rows(ws, start_row=5, end_row=end_row, n_cols=2)

    # 값 컬럼(B) 숫자 형식 적용
    _set_number_format(ws, start_row=5, end_row=end_row, col=2)

    # 틀 고정: 헤더(4행)까지 고정 → 스크롤 시 항목명 항상 표시
    ws.freeze_panes = "A5"

    # 컬럼 너비 자동 조정
    _auto_col_width(ws)


# =============================================================================
# [함수 3] write_maps_sheet
# =============================================================================

def write_maps_sheet(
    ws,
    figures: dict[str, go.Figure],
    img_refs: list,   # BytesIO 객체 수명 유지용 리스트 (호출자가 제공)
) -> None:
    """
    웨이퍼 맵 Figure를 PNG로 변환하여 2×2 그리드로 워크시트에 삽입.

    [2×2 그리드 배치]
    ┌─────────────────┬─────────────────┐
    │ A1  : Heatmap   │ I1  : Contour   │
    ├─────────────────┼─────────────────┤
    │ A31 : Line Scan │ I31 : 3D Surface│
    └─────────────────┴─────────────────┘
    행 오프셋 30 = 이미지 높이(350px) ÷ 엑셀 기본 행 높이(~13.5pt) ≈ 26행 + 여유 4행
    열 오프셋 I = 9번째 열 = 이미지 폭(400px) ÷ 엑셀 기본 열 너비(~8px) ≈ 8열 + 여유 1열

    [kaleido 없을 때 graceful degradation]
    PNG 변환 실패 시: 해당 셀 위치에 "이미지 생성 불가 (kaleido 미설치)" 텍스트 삽입.
    보고서가 이미지 없이 완성되어 사용자에게 다운로드 가능한 상태 유지.

    [img_refs 수명 관리]
    XLImage(BytesIO) 생성 후 wb.save() 시점에 BytesIO를 실제로 읽음.
    → 함수 내에서 생성한 BytesIO가 GC되면 저장 실패 가능성.
    → 호출자(generate_excel_report)가 제공한 img_refs 리스트에 추가해
       generate_excel_report 스코프 동안 BytesIO 강제 유지.

    인자:
        ws      : openpyxl 워크시트 객체
        figures : {"Heatmap": fig, "Contour": fig, "Line Scan": fig, "3D Surface": fig}
        img_refs: BytesIO 수명 유지용 리스트 (generate_excel_report에서 전달)
    """
    # 시트 제목
    ws["A1"] = "웨이퍼 맵 이미지"
    ws["A1"].font      = _TITLE_FONT
    ws["A1"].alignment = _LEFT_ALIGN

    # kaleido 미설치 경고 (전역 플래그로 확인)
    if not _KALEIDO_OK:
        ws["A2"] = (
            "⚠️ kaleido 미설치로 이미지 생성 불가. "
            "'pip install kaleido' 설치 후 재시도하세요."
        )
        ws["A2"].font = _WARN_FONT
        ws.merge_cells("A2:P2")
        return  # 이미지 없이 시트만 생성

    # ── 2×2 그리드 배치 설정 ──────────────────────────────────────────────────
    # (시트 이름, 셀 위치, 제목 행) 매핑
    grid_positions = [
        ("Heatmap",    "A3",  2),   # 1열 1행: A열 3행부터
        ("Contour",    "I3",  2),   # 2열 1행: I열 3행부터 (A+8)
        ("Line Scan",  "A33", 32),  # 1열 2행: A열 33행부터 (3+30)
        ("3D Surface", "I33", 32),  # 2열 2행: I열 33행부터
    ]

    # ── 각 Figure PNG 변환 및 삽입 ────────────────────────────────────────────
    for fig_name, cell_addr, title_row in grid_positions:
        fig = figures.get(fig_name)

        # 그림 제목 텍스트 (이미지 위에 표시)
        title_col = cell_addr[0]   # "A" 또는 "I"
        title_col_idx = ord(title_col) - ord("A") + 1
        ws.cell(row=title_row, column=title_col_idx, value=fig_name)
        ws.cell(row=title_row, column=title_col_idx).font = _BOLD_FONT

        if fig is None:
            # 해당 Figure가 전달되지 않은 경우
            ws.cell(
                row=title_row + 1,
                column=title_col_idx,
                value="(차트 없음)",
            ).font = _META_FONT
            continue

        # PNG 변환
        png_bytes = safe_fig_to_png(fig, width=600, height=500)

        if png_bytes is None:
            # kaleido 개별 변환 실패 시 텍스트 안내
            ws.cell(
                row=title_row + 1,
                column=title_col_idx,
                value="이미지 생성 불가 (kaleido 미설치 또는 변환 오류)",
            ).font = _WARN_FONT
            continue

        # ── BytesIO 생성 + img_refs에 추가 (수명 유지) ─────────────────────
        # wb.save() 시 XLImage가 BytesIO를 읽으므로
        # 이 BytesIO는 generate_excel_report가 반환될 때까지 살아있어야 함
        img_io = io.BytesIO(png_bytes)
        img_refs.append(img_io)   # ★ 수명 유지를 위한 참조 추가

        # ── XLImage 생성 및 워크시트 삽입 ──────────────────────────────────
        xl_img        = XLImage(img_io)
        xl_img.width  = 400    # 픽셀 단위 (엑셀 내 표시 너비)
        xl_img.height = 350    # 픽셀 단위 (엑셀 내 표시 높이)
        ws.add_image(xl_img, cell_addr)


# =============================================================================
# [함수 4] generate_excel_report
# =============================================================================

def generate_excel_report(
    filename: str,
    stats: dict,
    df_display: pd.DataFrame,
    fig_heatmap: go.Figure,
    fig_contour: go.Figure,
    fig_linescan: go.Figure,
    fig_3d: go.Figure,
    include_maps: bool = True,
    include_raw: bool = True,
    max_raw_rows: int = 5000,
    gpc_data: dict | None = None,   # {"stats": dict, "fig": go.Figure}
) -> bytes:
    """
    분석 결과를 다중 시트 xlsx 파일로 생성하여 바이너리 반환.

    [@st.cache_data 미적용 이유]
    go.Figure 인자들이 mutable → hash() 불가 → 일반 함수로 구현.
    버튼 클릭 시에만 호출되므로 성능 영향 없음.

    [시트 구성]
    "요약"       : 파일 정보 + 주요 통계 지표 (항상 포함)
    "웨이퍼 맵" : 4종 차트 PNG 이미지 2×2 그리드 (include_maps=True 시)
    "원시 데이터": 측정 데이터 테이블 (include_raw=True 시, 최대 max_raw_rows행)
    "GPC 분석"  : GPC 통계 + GPC 맵 이미지 (gpc_data 전달 시)

    [BytesIO 수명 관리]
    _img_refs 리스트로 모든 이미지 BytesIO를 이 함수 스코프 동안 유지.
    wb.save(buf) 시점에 XLImage가 BytesIO를 읽으므로 이 시점까지 살아있어야 함.
    → 함수 반환 후 _img_refs GC → BytesIO 해제 (정상).

    인자:
        filename    : 보고서 파일명 (요약 시트 헤더에 표시)
        stats       : calculate_stats() 반환 딕셔너리
        df_display  : 원시 데이터 DataFrame (x, y, data 컬럼)
        fig_heatmap : 2D Heatmap Figure
        fig_contour : Contour Map Figure
        fig_linescan: Line Scan Figure
        fig_3d      : 3D Surface Figure
        include_maps: True이면 "웨이퍼 맵" 시트 생성
        include_raw : True이면 "원시 데이터" 시트 생성
        max_raw_rows: 원시 데이터 최대 행 수 (기본 5000)
        gpc_data    : GPC 분석 데이터 dict 또는 None
                     {"stats": dict, "fig": go.Figure}

    반환:
        bytes: xlsx 파일 바이너리 (st.download_button에 직접 전달 가능)
    """
    # BytesIO 수명 유지용 리스트 (wb.save() 시점까지 GC 방지)
    _img_refs: list = []

    # ── Workbook 생성 ─────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)   # 기본 생성되는 빈 Sheet 제거

    # 보고서 생성 시각 (모든 시트 메타에 사용)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── 시트 1: 요약 (항상 생성) ──────────────────────────────────────────────
    ws_summary = wb.create_sheet("요약")

    # 메인 제목 (A1:F1 병합)
    ws_summary["A1"] = "웨이퍼 맵 분석 보고서"
    ws_summary["A1"].font      = _TITLE_FONT
    ws_summary["A1"].alignment = _CENTER_ALIGN
    ws_summary.merge_cells("A1:F1")
    ws_summary.row_dimensions[1].height = 30

    # 메타 정보 (파일명, 생성 시각)
    ws_summary["A2"] = f"분석 파일: {filename}"
    ws_summary["A2"].font      = _META_FONT
    ws_summary["A2"].alignment = _LEFT_ALIGN
    ws_summary.merge_cells("A2:F2")

    ws_summary["A3"] = f"보고서 생성: {now_str}"
    ws_summary["A3"].font      = _META_FONT
    ws_summary["A3"].alignment = _LEFT_ALIGN
    ws_summary.merge_cells("A3:F3")

    # 포함 내용 요약
    included = []
    if include_maps:  included.append("웨이퍼 맵 이미지")
    if include_raw:   included.append("원시 데이터")
    if gpc_data:      included.append("GPC 분석")
    ws_summary["A4"] = f"포함 내용: {', '.join(included) if included else '통계만'}"
    ws_summary["A4"].font = _META_FONT
    ws_summary.merge_cells("A4:F4")

    # 통계 헤더 (Row 6)
    stat_headers = ["항목", "값", "", "설명"]
    # 주요 통계 설명 딕셔너리
    stat_descriptions = {
        "Mean":           "산술 평균 (전체 측정 포인트)",
        "Maximum":        "최대값",
        "Minimum":        "최솟값",
        "Std Dev":        "표준편차",
        "Uniformity (%)": "균일도 = σ/μ × 100 (낮을수록 우수)",
        "Range":          "최대 - 최소",
        "No. Sites":      "유효 측정 포인트 수",
    }

    for col_idx, h in enumerate(["항목", "값", "", "설명"], start=1):
        ws_summary.cell(row=6, column=col_idx, value=h if h else "")
    _style_header_row(ws_summary, row=6, n_cols=2)   # 항목, 값만 스타일
    ws_summary.cell(row=6, column=4, value="설명").font = _HEADER_FONT
    ws_summary.cell(row=6, column=4).fill      = _HEADER_FILL
    ws_summary.cell(row=6, column=4).alignment = _CENTER_ALIGN
    ws_summary.cell(row=6, column=4).border    = _THIN_BORDER

    # 통계 데이터 행 (Row 7~)
    for row_offset, (key, val) in enumerate(stats.items()):
        row = 7 + row_offset
        ws_summary.cell(row=row, column=1, value=key)
        ws_summary.cell(row=row, column=2, value=val)
        ws_summary.cell(row=row, column=4,
                        value=stat_descriptions.get(key, ""))

    end_stat_row = 6 + len(stats)
    _style_data_rows(ws_summary, start_row=7, end_row=end_stat_row, n_cols=2)
    # 설명 컬럼도 스타일 적용
    for row in range(7, end_stat_row + 1):
        cell = ws_summary.cell(row=row, column=4)
        cell.font      = _BODY_FONT
        cell.border    = _THIN_BORDER
        cell.alignment = _LEFT_ALIGN
        cell.fill      = _ALT_FILL if row % 2 == 0 else _WHITE_FILL

    # 값 컬럼 숫자 형식
    _set_number_format(ws_summary, start_row=7, end_row=end_stat_row, col=2)

    ws_summary.freeze_panes = "A7"
    ws_summary.column_dimensions["A"].width = 18
    ws_summary.column_dimensions["B"].width = 16
    ws_summary.column_dimensions["C"].width = 3
    ws_summary.column_dimensions["D"].width = 40

    # ── 시트 2: 상세 통계 (write_stats_sheet 활용) ───────────────────────────
    ws_stats = wb.create_sheet("통계 상세")
    write_stats_sheet(ws_stats, stats, filename, now_str)

    # ── 시트 3: 웨이퍼 맵 이미지 (선택적) ───────────────────────────────────
    if include_maps:
        ws_maps = wb.create_sheet("웨이퍼 맵")
        figures = {
            "Heatmap":    fig_heatmap,
            "Contour":    fig_contour,
            "Line Scan":  fig_linescan,
            "3D Surface": fig_3d,
        }
        # ★ _img_refs를 전달 → write_maps_sheet 내부에서 BytesIO를 추가
        #   → wb.save() 시점까지 BytesIO 수명 유지 보장
        write_maps_sheet(ws_maps, figures, _img_refs)

        # 이미지 셀 크기에 맞게 행 높이 조정
        for row_num in range(1, 65):
            ws_maps.row_dimensions[row_num].height = 15

    # ── 시트 4: 원시 데이터 (선택적) ────────────────────────────────────────
    if include_raw and df_display is not None and not df_display.empty:
        ws_raw = wb.create_sheet("원시 데이터")

        # 제목
        ws_raw["A1"] = f"원시 데이터: {filename}"
        ws_raw["A1"].font      = _TITLE_FONT
        ws_raw["A1"].alignment = _LEFT_ALIGN
        n_display_cols = len(df_display.columns)
        ws_raw.merge_cells(f"A1:{get_column_letter(n_display_cols)}1")

        ws_raw["A2"] = f"생성: {now_str}"
        ws_raw["A2"].font = _META_FONT

        # 컬럼 헤더 (Row 4)
        for col_idx, col_name in enumerate(df_display.columns, start=1):
            ws_raw.cell(row=4, column=col_idx, value=col_name)
        _style_header_row(ws_raw, row=4, n_cols=n_display_cols)

        # 데이터 행 (최대 max_raw_rows 행으로 제한)
        df_out   = df_display.head(max_raw_rows)
        n_out    = len(df_out)

        for row_offset, (_, data_row) in enumerate(df_out.iterrows()):
            row = 5 + row_offset
            for col_idx, val in enumerate(data_row, start=1):
                cell = ws_raw.cell(row=row, column=col_idx, value=val)
                # 숫자 컬럼 형식 적용
                if isinstance(val, float):
                    cell.number_format = _NUM_FORMAT

        end_raw_row = 4 + n_out
        _style_data_rows(ws_raw, start_row=5, end_row=end_raw_row,
                          n_cols=n_display_cols, first_col_left=False)

        ws_raw.freeze_panes = "A5"

        # 행 수 제한 초과 경고
        if len(df_display) > max_raw_rows:
            warn_row = end_raw_row + 2
            ws_raw.cell(
                row=warn_row, column=1,
                value=(f"* 표시 제한: {max_raw_rows:,}행 "
                       f"(전체 {len(df_display):,}행). "
                       f"전체 데이터는 앱에서 CSV로 다운로드하세요."),
            ).font = _WARN_FONT
            ws_raw.merge_cells(
                f"A{warn_row}:{get_column_letter(n_display_cols)}{warn_row}"
            )

        _auto_col_width(ws_raw)

    # ── 시트 5: GPC 분석 (선택적) ────────────────────────────────────────────
    if gpc_data is not None:
        ws_gpc = wb.create_sheet("GPC 분석")

        # GPC 통계
        gpc_stats = gpc_data.get("stats", {})
        gpc_fig   = gpc_data.get("fig", None)

        ws_gpc["A1"] = "GPC (Growth Per Cycle) 분석"
        ws_gpc["A1"].font      = _TITLE_FONT
        ws_gpc["A1"].alignment = _LEFT_ALIGN
        ws_gpc.merge_cells("A1:C1")

        ws_gpc["A2"] = f"분석 파일: {filename}    생성: {now_str}"
        ws_gpc["A2"].font = _META_FONT

        # GPC 통계 테이블
        gpc_stat_headers = ["항목", "값"]
        for col_idx, h in enumerate(gpc_stat_headers, start=1):
            ws_gpc.cell(row=4, column=col_idx, value=h)
        _style_header_row(ws_gpc, row=4, n_cols=2)

        for row_offset, (key, val) in enumerate(gpc_stats.items()):
            row = 5 + row_offset
            ws_gpc.cell(row=row, column=1, value=key)
            ws_gpc.cell(row=row, column=2, value=val)

        end_gpc_row = 4 + len(gpc_stats)
        _style_data_rows(ws_gpc, start_row=5, end_row=end_gpc_row, n_cols=2)
        _set_number_format(ws_gpc, start_row=5, end_row=end_gpc_row, col=2)
        _auto_col_width(ws_gpc)

        # GPC Figure 이미지 삽입 (kaleido 가용 시)
        if gpc_fig is not None:
            gpc_png = safe_fig_to_png(gpc_fig, width=600, height=500)
            if gpc_png is not None:
                gpc_img_io = io.BytesIO(gpc_png)
                _img_refs.append(gpc_img_io)   # 수명 유지
                xl_img = XLImage(gpc_img_io)
                xl_img.width  = 500
                xl_img.height = 420
                # GPC 통계 테이블 아래에 이미지 배치
                img_start_row = end_gpc_row + 3
                ws_gpc.add_image(xl_img, f"A{img_start_row}")
            else:
                warn_row = end_gpc_row + 3
                ws_gpc.cell(
                    row=warn_row, column=1,
                    value="GPC 차트 이미지 생성 불가 (kaleido 미설치)",
                ).font = _WARN_FONT

    # ── xlsx 직렬화 ───────────────────────────────────────────────────────────
    # ★ 이 시점에 _img_refs의 모든 BytesIO가 살아있어야 XLImage가 정상 저장됨
    buf = io.BytesIO()
    wb.save(buf)
    # wb.save() 완료 후 _img_refs는 이 함수 스코프에서 계속 유지
    # → 함수 반환 후 GC (안전)

    return buf.getvalue()


# =============================================================================
# [함수 5] render_report_tab (UI 렌더러)
# =============================================================================

def render_report_tab(
    filename: str,
    stats: dict,
    df_display: pd.DataFrame,
    fig_heatmap: go.Figure,
    fig_contour: go.Figure,
    fig_linescan: go.Figure,
    fig_3d: go.Figure,
    gpc_data: dict | None,
) -> None:
    """
    Excel 보고서 생성 탭의 전체 UI를 렌더링.

    [레이아웃 구조]
    ┌─────────────────────────────────────────────────────────────────────────┐
    │ kaleido 미설치 경고 (필요 시만 표시)                                      │
    ├──────────────────────────────────────────┬──────────────────────────────┤
    │  포함 내용 옵션 (체크박스)                  │  원시 데이터 행 수 슬라이더    │
    │  [✅] 웨이퍼 맵 이미지 (kaleido 필요)        │  (include_raw=True 시만)      │
    │  [✅] 원시 데이터                            │                              │
    │  [□ ] GPC 분석 (gpc_data 있을 때만)         │                              │
    ├──────────────────────────────────────────┴──────────────────────────────┤
    │  📊 통계 미리보기 (st.dataframe)                                          │
    ├─────────────────────────────────────────────────────────────────────────┤
    │  [📥 Excel 보고서 생성 및 다운로드] 버튼                                   │
    │    → st.spinner("보고서 생성 중...") 동안 generate_excel_report() 실행     │
    │    → 완료 시 st.download_button() 표시                                    │
    └─────────────────────────────────────────────────────────────────────────┘

    [중복 클릭 방지]
    rep_generating session_state 플래그:
      True  → 버튼 비활성화 (disabled=True)
      False → 버튼 활성화 (기본)
    생성 완료 후 rep_bytes에 결과 저장 → download_button 표시 유지.

    [보고서 생성 결과 보존]
    session_state["rep_bytes"]에 xlsx bytes 저장.
    → st.rerun() 후에도 download_button 표시 가능.
    → 새 보고서 생성 버튼 클릭 시 rep_bytes 초기화.

    인자:
        filename    : 현재 분석 파일명
        stats       : calculate_stats() 반환 딕셔너리
        df_display  : 표준화된 측정 DataFrame (x, y, data)
        fig_heatmap : 2D Heatmap Figure
        fig_contour : Contour Map Figure
        fig_linescan: Line Scan Figure
        fig_3d      : 3D Surface Figure
        gpc_data    : GPC 분석 데이터 또는 None
    """
    # ── session_state 초기화 ──────────────────────────────────────────────────
    if _SS_BYTES      not in st.session_state: st.session_state[_SS_BYTES]      = None
    if _SS_GENERATING not in st.session_state: st.session_state[_SS_GENERATING] = False

    # ── kaleido 미설치 경고 ───────────────────────────────────────────────────
    # safe_fig_to_png로 실제 변환 시도 → None 반환이면 미설치 확인
    if not _KALEIDO_OK:
        st.warning(
            "⚠️ **kaleido 미설치**: 웨이퍼 맵 이미지를 포함한 보고서 생성 불가. "
            "이미지 없이 통계 데이터만 포함된 보고서가 생성됩니다.\n\n"
            "이미지 포함 보고서가 필요하면: `pip install kaleido`"
        )

    # ── 포함 내용 옵션 + 원시 데이터 행 수 ──────────────────────────────────
    opt_col, ctrl_col = st.columns([2, 1])

    with opt_col:
        st.markdown("##### 📋 보고서 포함 내용")

        # 웨이퍼 맵 이미지 포함 여부 (kaleido 없으면 비활성화)
        include_maps: bool = st.checkbox(
            "🗺️ 웨이퍼 맵 이미지 (kaleido 필요)",
            value=st.session_state.get(_SS_INC_MAPS, True),
            key=_SS_INC_MAPS,
            disabled=not _KALEIDO_OK,   # kaleido 없으면 선택 불가
            help=(
                "Heatmap, Contour, Line Scan, 3D Surface 이미지를 보고서에 포함합니다.\n"
                "kaleido 설치 필요: `pip install kaleido`"
            ),
        )

        # 원시 데이터 포함 여부
        include_raw: bool = st.checkbox(
            "📊 원시 데이터 (x, y, data 테이블)",
            value=st.session_state.get(_SS_INC_RAW, True),
            key=_SS_INC_RAW,
            help="측정 데이터 테이블을 별도 시트로 포함합니다.",
        )

        # GPC 분석 포함 여부 (gpc_data 없으면 비활성화)
        include_gpc: bool = st.checkbox(
            "⚙️ GPC 분석 결과",
            value=st.session_state.get(_SS_INC_GPC, gpc_data is not None),
            key=_SS_INC_GPC,
            disabled=(gpc_data is None),   # GPC 데이터 없으면 선택 불가
            help=(
                "GPC 탭에서 계산된 GPC 통계와 차트를 포함합니다.\n"
                "GPC 탭에서 먼저 계산을 실행해야 합니다."
                if gpc_data is None
                else "GPC 분석 결과를 별도 시트로 포함합니다."
            ),
        )

    with ctrl_col:
        # 원시 데이터 행 수 제한 슬라이더 (include_raw=True 시만 표시)
        if include_raw:
            st.markdown("##### ⚙️ 원시 데이터 설정")
            max_raw_rows: int = st.slider(
                "최대 행 수",
                min_value=100,
                max_value=10000,
                value=st.session_state.get(_SS_NROWS, 5000),
                step=100,
                key=_SS_NROWS,
                help=(
                    "원시 데이터 시트에 포함할 최대 행 수.\n"
                    "행 수가 많을수록 파일 크기가 커집니다.\n"
                    "전체 데이터는 메인 화면 CSV 다운로드를 사용하세요."
                ),
            )
            total_rows = len(df_display) if df_display is not None else 0
            if total_rows > max_raw_rows:
                st.info(
                    f"전체 {total_rows:,}행 중 {max_raw_rows:,}행만 포함됩니다."
                )
        else:
            max_raw_rows = 5000   # include_raw=False이면 사용하지 않지만 기본값 설정

    st.markdown("---")

    # ── 통계 미리보기 ────────────────────────────────────────────────────────
    st.markdown("##### 📊 통계 미리보기 (보고서 '요약' 시트 내용)")
    preview_df = pd.DataFrame([
        {"항목": k, "값": v,
         "단위": "%" if "Uniformity" in k else ("개" if "Sites" in k else "")}
        for k, v in stats.items()
    ])
    st.dataframe(preview_df, use_container_width=True, hide_index=True)

    st.markdown("---")

    # ── 보고서 생성 버튼 ─────────────────────────────────────────────────────
    # 중복 클릭 방지: 생성 중이면 버튼 비활성화
    is_generating = st.session_state.get(_SS_GENERATING, False)

    btn_label = "⏳ 보고서 생성 중..." if is_generating else "📥 Excel 보고서 생성 및 다운로드"
    generate_clicked = st.button(
        btn_label,
        type="primary",
        key="rep_gen_btn",
        disabled=is_generating,
        use_container_width=True,
        help="클릭하면 Excel 보고서를 생성하고 즉시 다운로드합니다.",
    )

    if generate_clicked:
        # ── 보고서 생성 실행 ──────────────────────────────────────────────────
        st.session_state[_SS_GENERATING] = True
        st.session_state[_SS_BYTES]      = None   # 이전 보고서 초기화

        with st.spinner("보고서 생성 중... 잠시 기다려주세요."):
            try:
                xl_bytes = generate_excel_report(
                    filename=filename,
                    stats=stats,
                    df_display=df_display,
                    fig_heatmap=fig_heatmap,
                    fig_contour=fig_contour,
                    fig_linescan=fig_linescan,
                    fig_3d=fig_3d,
                    include_maps=include_maps and _KALEIDO_OK,
                    include_raw=include_raw,
                    max_raw_rows=max_raw_rows,
                    gpc_data=gpc_data if include_gpc else None,
                )
                st.session_state[_SS_BYTES] = xl_bytes
                st.success("✅ 보고서 생성 완료! 아래 버튼으로 다운로드하세요.")

            except Exception as e:
                st.error(
                    f"❌ 보고서 생성 실패: {type(e).__name__}: {e}\n\n"
                    "문제가 지속되면 '웨이퍼 맵 이미지' 옵션을 해제하고 재시도하세요."
                )

            finally:
                # 성공/실패 모두 생성 플래그 해제 → 버튼 재활성화
                st.session_state[_SS_GENERATING] = False

    # ── 다운로드 버튼 (보고서 생성 완료 시) ──────────────────────────────────
    xl_bytes = st.session_state.get(_SS_BYTES)

    if xl_bytes is not None:
        # 파일명: 원본 파일명에서 확장자 제거 + 타임스탬프 추가
        base_name   = os.path.splitext(filename)[0][:30]   # 30자로 제한
        ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_name = f"{base_name}_report_{ts}.xlsx"

        st.download_button(
            label="⬇️ Excel 파일 다운로드",
            data=xl_bytes,              # ★ bytes 타입 전달 (BytesIO.getvalue() 반환값)
            file_name=report_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary",
            use_container_width=True,
            key="rep_download_btn",
            help=f"'{report_name}' 파일로 다운로드됩니다.",
        )

        # 파일 크기 정보 표시
        size_kb = len(xl_bytes) / 1024
        size_str = f"{size_kb:.1f} KB" if size_kb < 1024 else f"{size_kb/1024:.2f} MB"
        st.caption(
            f"📄 파일명: `{report_name}`  |  크기: {size_str}  |  "
            f"포함: "
            + (", ".join(filter(None, [
                "요약·통계",
                "웨이퍼 맵" if include_maps and _KALEIDO_OK else None,
                "원시 데이터" if include_raw else None,
                "GPC 분석" if include_gpc and gpc_data else None,
            ])))
        )